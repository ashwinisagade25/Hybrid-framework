import openpyxl

def getRowCount(file,sheet):
    wbook=openpyxl.load_workbook(file)
    sheet=wbook.active
    return sheet.max_row

def getColumnCount(file,sheet):
    wbook=openpyxl.load_workbook(file)
    sheet=wbook.active
    return sheet.max_col

def readData(file,sheet,rown,coln):
    wbook=openpyxl.load_workbook(file)
    sheet=wbook.active
    return sheet.cell(row=rown,column=coln).value

def writeData(file,sheet,rown,coln,data):
    wbook=openpyxl.load_workbook(file)
    sheet=wbook.active
    sheet.cell(row=rown,column=coln).value = data
    wbook.save(file)

